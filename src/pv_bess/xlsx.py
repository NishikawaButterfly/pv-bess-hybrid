"""Optional Excel export of a completed run's published artifact pair.

The workbook re-reads ``summary.json`` and ``dispatch.csv`` exactly as
``pv-bess run`` published them: a Summary sheet with KPIs, solver metadata,
provenance hashes, and the limitations list; a Cash flows sheet with the
year-by-year series; and a Dispatch sheet with the per-interval columns.
Formatting stays sober -- bold headers, number formats, frozen header rows,
and column widths. Charts and semantic colors are deliberately absent.
"""

from __future__ import annotations

import csv
import json
import os
import tempfile
from collections.abc import Sequence
from contextlib import suppress
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from pv_bess.io import _MAX_INTERVALS, ScenarioFileError, _read_bounded_text

_CREATOR = "pv-bess-hybrid"
_MAX_SUMMARY_BYTES = 5_000_000
_MAX_DISPATCH_BYTES = 100_000_000

_EUR_FORMAT = "#,##0.00"
_ENERGY_FORMAT = "0.000"
_POWER_FORMAT = "#,##0.0"
_FRACTION_FORMAT = "0.0000"
_SCALAR_FORMAT = "0.00"
_GAP_FORMAT = "0.00E+00"

_BOLD_FONT = Font(bold=True)

_DISPATCH_COLUMNS: tuple[tuple[str, str, str | None], ...] = (
    ("timestamp", "Timestamp", None),
    ("interval_hours", "Interval (h)", _SCALAR_FORMAT),
    ("pv_power_kw", "PV power (kW)", _POWER_FORMAT),
    ("market_price_eur_per_mwh", "Market price (EUR/MWh)", _EUR_FORMAT),
    ("pv_export_kw", "PV export (kW)", _POWER_FORMAT),
    ("pv_charge_kw", "PV charge (kW)", _POWER_FORMAT),
    ("grid_charge_kw", "Grid charge (kW)", _POWER_FORMAT),
    ("battery_export_kw", "Battery export (kW)", _POWER_FORMAT),
    ("grid_export_kw", "Grid export (kW)", _POWER_FORMAT),
    ("curtailed_pv_kw", "Curtailed PV (kW)", _POWER_FORMAT),
    ("soc_start_kwh", "SOC start (kWh DC)", _POWER_FORMAT),
    ("soc_end_kwh", "SOC end (kWh DC)", _POWER_FORMAT),
    ("market_value_eur", "Market value (EUR)", _EUR_FORMAT),
    ("degradation_cost_eur", "Degradation cost (EUR)", _EUR_FORMAT),
    ("net_operating_value_eur", "Net operating value (EUR)", _EUR_FORMAT),
)

_DISPATCH_CSV_COLUMNS: tuple[str, ...] = (
    "dispatch_input_sha256",
    "analysis_input_sha256",
    *(field for field, _, _ in _DISPATCH_COLUMNS),
)

_DISPATCH_KPI_ROWS: tuple[tuple[str, str, str, str], ...] = (
    ("pv_energy_mwh", "PV energy", "MWh", _ENERGY_FORMAT),
    ("export_energy_mwh", "Export energy", "MWh", _ENERGY_FORMAT),
    ("grid_charge_energy_mwh", "Grid charge energy", "MWh", _ENERGY_FORMAT),
    ("battery_discharge_energy_mwh", "Battery discharge energy", "MWh", _ENERGY_FORMAT),
    ("curtailed_energy_mwh", "Curtailed energy", "MWh", _ENERGY_FORMAT),
    ("baseline_export_energy_mwh", "Baseline export energy", "MWh", _ENERGY_FORMAT),
    ("baseline_curtailed_energy_mwh", "Baseline curtailed energy", "MWh", _ENERGY_FORMAT),
    ("market_value_eur", "Market value", "EUR", _EUR_FORMAT),
    ("degradation_cost_eur", "Degradation cost", "EUR", _EUR_FORMAT),
    ("operating_value_eur", "Operating value", "EUR", _EUR_FORMAT),
    ("baseline_market_value_eur", "Baseline market value", "EUR", _EUR_FORMAT),
    ("incremental_operating_value_eur", "Incremental operating value", "EUR", _EUR_FORMAT),
    ("equivalent_full_cycles", "Equivalent full cycles", "cycles", _SCALAR_FORMAT),
    ("curtailment_reduction_fraction", "Curtailment reduction", "fraction", _FRACTION_FORMAT),
    ("ending_soc_kwh", "Ending SOC", "kWh DC", _POWER_FORMAT),
)

_FINANCIAL_KPI_ROWS: tuple[tuple[str, str, str, str], ...] = (
    ("npv_eur", "Net present value", "EUR", _EUR_FORMAT),
    ("irr_fraction", "Internal rate of return", "fraction", _FRACTION_FORMAT),
    ("simple_payback_years", "Simple payback", "years", _SCALAR_FORMAT),
    ("discounted_payback_years", "Discounted payback", "years", _SCALAR_FORMAT),
    ("lcos_eur_per_mwh", "Levelized cost of storage", "EUR/MWh", _EUR_FORMAT),
)


class ResultsFileError(ValueError):
    """Raised when a run directory is unreadable, malformed, or inconsistent."""


def _read_results_text(path: Path, maximum_bytes: int) -> str:
    try:
        return _read_bounded_text(path, maximum_bytes)
    except ScenarioFileError as exc:
        raise ResultsFileError(str(exc)) from exc


def load_run_directory(directory: str | Path) -> tuple[dict[str, Any], list[dict[str, str]]]:
    """Load and cross-check the ``summary.json``/``dispatch.csv`` pair of one run."""

    run_directory = Path(directory).resolve()
    summary_text = _read_results_text(run_directory / "summary.json", _MAX_SUMMARY_BYTES)
    try:
        summary = json.loads(summary_text)
    except json.JSONDecodeError as exc:
        raise ResultsFileError(f"invalid summary JSON: {exc}") from exc
    if not isinstance(summary, dict):
        raise ResultsFileError("summary.json must contain a JSON object")
    try:
        dispatch_hash = summary["dispatch_input_sha256"]
        analysis_hash = summary["analysis_input_sha256"]
    except KeyError as exc:
        raise ResultsFileError(f"summary.json is missing required key {exc}") from exc

    dispatch_text = _read_results_text(run_directory / "dispatch.csv", _MAX_DISPATCH_BYTES)
    reader = csv.DictReader(dispatch_text.splitlines())
    if reader.fieldnames is None or tuple(reader.fieldnames) != _DISPATCH_CSV_COLUMNS:
        raise ResultsFileError(
            "dispatch.csv header must be exactly: " + ",".join(_DISPATCH_CSV_COLUMNS)
        )

    rows: list[dict[str, str]] = []
    for line_number, row in enumerate(reader, start=2):
        if len(rows) >= _MAX_INTERVALS:
            raise ResultsFileError(f"dispatch.csv exceeds the {_MAX_INTERVALS}-row limit")
        if None in row or None in row.values():
            raise ResultsFileError(f"invalid dispatch.csv row {line_number}: field count mismatch")
        if row["dispatch_input_sha256"] != dispatch_hash:
            raise ResultsFileError(
                f"dispatch.csv row {line_number} does not match the summary.json provenance hashes"
            )
        if row["analysis_input_sha256"] != analysis_hash:
            raise ResultsFileError(
                f"dispatch.csv row {line_number} does not match the summary.json provenance hashes"
            )
        rows.append(row)
    if not rows:
        raise ResultsFileError("dispatch.csv contains no interval rows")
    return summary, rows


class _SheetWriter:
    """Append rows through an explicit cursor so spacer rows stay deterministic."""

    def __init__(self, sheet: Any) -> None:
        self._sheet = sheet
        self._row = 0

    def next_row(self) -> int:
        self._row += 1
        return self._row

    def header(self, *labels: str) -> None:
        row = self.next_row()
        for column, label in enumerate(labels, start=1):
            self._sheet.cell(row=row, column=column, value=label).font = _BOLD_FONT
        self._sheet.freeze_panes = "A2"

    def section(self, title: str) -> None:
        self.next_row()
        self._sheet.cell(row=self.next_row(), column=1, value=title).font = _BOLD_FONT

    def note(self, text: str) -> None:
        self._sheet.cell(row=self.next_row(), column=1, value=text)

    def item(
        self,
        label: str,
        value: Any,
        unit: str = "",
        number_format: str | None = None,
    ) -> None:
        row = self.next_row()
        self._sheet.cell(row=row, column=1, value=label)
        if value is None:
            self._sheet.cell(row=row, column=2, value="n/a")
        else:
            cell = self._sheet.cell(row=row, column=2, value=value)
            if number_format is not None:
                cell.number_format = number_format
        if unit:
            self._sheet.cell(row=row, column=3, value=unit)


def _set_column_widths(sheet: Any, widths: Sequence[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _populate_summary_sheet(sheet: Any, summary: dict[str, Any]) -> None:
    sheet.title = "Summary"
    _set_column_widths(sheet, (34, 70, 12))
    writer = _SheetWriter(sheet)
    writer.header("Item", "Value", "Unit")

    writer.section("Run")
    writer.item("Scenario name", summary["scenario_name"])
    writer.item("Generated at (UTC)", summary["generated_at_utc"])
    writer.item("Schema version", summary["schema_version"])
    writer.item("Model version", summary["model_version"])

    writer.section("Provenance")
    writer.item("Dispatch input SHA-256", summary["dispatch_input_sha256"])
    writer.item("Analysis input SHA-256", summary["analysis_input_sha256"])

    solver = summary["solver"]
    writer.section("Solver")
    writer.item("Interface", f"{solver['interface_name']} {solver['interface_version']}")
    writer.item("Backend", f"{solver['backend_name']} {solver['backend_version'] or 'unknown'}")
    writer.item(
        "Requested relative MIP gap",
        solver["requested_relative_mip_gap"],
        number_format=_GAP_FORMAT,
    )
    writer.item(
        "Time limit per phase",
        solver["time_limit_seconds_per_phase"],
        "s",
        _SCALAR_FORMAT,
    )
    for key, label in (("economic", "Economic phase"), ("refinement", "Refinement phase")):
        phase = solver["phases"][key]
        writer.item(f"{label} status", phase["status"])
        writer.item(
            f"{label} objective", phase["objective_value"], phase["objective_unit"], _EUR_FORMAT
        )
        writer.item(
            f"{label} achieved MIP gap",
            phase["achieved_relative_mip_gap"],
            number_format=_GAP_FORMAT,
        )

    dispatch_summary = summary["dispatch_summary"]
    writer.section("Dispatch KPIs")
    for key, label, unit, number_format in _DISPATCH_KPI_ROWS:
        writer.item(label, dispatch_summary[key], unit, number_format)

    financial_summary = summary["financial_summary"]
    writer.section("Financial KPIs")
    for key, label, unit, number_format in _FINANCIAL_KPI_ROWS:
        writer.item(label, financial_summary[key], unit, number_format)

    # Present only when the run modeled multi-year capacity fade; older
    # summaries and zero-fade runs have no capacity_fade block.
    capacity_fade = financial_summary.get("capacity_fade")
    if isinstance(capacity_fade, dict):
        fractions = capacity_fade["capacity_fraction_by_year"]
        if not isinstance(fractions, list) or not fractions:
            raise ValueError("capacity_fraction_by_year must be a non-empty array")
        writer.section("Capacity fade")
        writer.item(
            "Calendar fade per year",
            capacity_fade["calendar_fade_fraction_per_year"],
            "fraction",
            _FRACTION_FORMAT,
        )
        writer.item(
            "Cycling fade per equivalent full cycle",
            capacity_fade["cycling_fade_fraction_per_efc"],
            "fraction",
            _FRACTION_FORMAT,
        )
        writer.item(
            "Minimum capacity fraction",
            capacity_fade["minimum_capacity_fraction"],
            "fraction",
            _FRACTION_FORMAT,
        )
        writer.item(
            "Annualized equivalent full cycles",
            capacity_fade["annualized_equivalent_full_cycles"],
            "cycles/year",
            _SCALAR_FORMAT,
        )
        writer.item("First-year capacity", fractions[0], "fraction", _FRACTION_FORMAT)
        writer.item(
            f"Final-year capacity (year {len(fractions)})",
            fractions[-1],
            "fraction",
            _FRACTION_FORMAT,
        )

    # Absent from summaries published before the warning channel existed, and
    # empty for a run whose assumptions raised nothing.
    warnings = financial_summary.get("warnings")
    if isinstance(warnings, list) and warnings:
        writer.section("Warnings")
        for warning in warnings:
            writer.note(str(warning))

    writer.section("Limitations")
    for limitation in summary["limitations"]:
        writer.note(str(limitation))


def _populate_cash_flow_sheet(sheet: Any, summary: dict[str, Any]) -> None:
    _set_column_widths(sheet, (8, 18))
    writer = _SheetWriter(sheet)
    writer.header("Year", "Cash flow (EUR)")
    for year, value in enumerate(summary["financial_summary"]["cash_flows_eur"]):
        row = writer.next_row()
        sheet.cell(row=row, column=1, value=year)
        sheet.cell(row=row, column=2, value=value).number_format = _EUR_FORMAT


def _float(row: dict[str, str], key: str, line_number: int) -> float:
    try:
        return float(row[key])
    except ValueError as exc:
        raise ResultsFileError(f"invalid dispatch.csv row {line_number}: {exc}") from exc


def _populate_dispatch_sheet(sheet: Any, dispatch_rows: Sequence[dict[str, str]]) -> None:
    _set_column_widths(
        sheet,
        (26, *(max(len(header) + 2, 12) for _, header, _ in _DISPATCH_COLUMNS[1:])),
    )
    writer = _SheetWriter(sheet)
    writer.header(*(header for _, header, _ in _DISPATCH_COLUMNS))
    for line_number, row in enumerate(dispatch_rows, start=2):
        excel_row = writer.next_row()
        sheet.cell(row=excel_row, column=1, value=row["timestamp"])
        for column, (field, _, number_format) in enumerate(_DISPATCH_COLUMNS[1:], start=2):
            cell = sheet.cell(row=excel_row, column=column, value=_float(row, field, line_number))
            cell.number_format = number_format


def build_workbook(summary: dict[str, Any], dispatch_rows: Sequence[dict[str, str]]) -> Any:
    """Return an ``openpyxl`` workbook rendering one completed run."""

    workbook = Workbook()
    try:
        _populate_summary_sheet(workbook.active, summary)
        _populate_cash_flow_sheet(workbook.create_sheet("Cash flows"), summary)
    except (AttributeError, KeyError, TypeError, ValueError) as exc:
        raise ResultsFileError(f"summary.json is not a complete run summary: {exc!r}") from exc
    _populate_dispatch_sheet(workbook.create_sheet("Dispatch"), dispatch_rows)
    # Keep the file metadata anonymous: the project name is the only identity.
    workbook.properties.creator = _CREATOR
    workbook.properties.lastModifiedBy = _CREATOR
    return workbook


def _publish_bytes(path: Path, content: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        dir=path.parent,
        prefix=f".{path.name}.",
        suffix=".stage.tmp",
    )
    temporary_path = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(content)
            handle.flush()
            os.fsync(handle.fileno())
        temporary_path.replace(path)
    finally:
        with suppress(OSError):
            temporary_path.unlink(missing_ok=True)


def export_workbook(
    input_directory: str | Path,
    output_path: str | Path,
    *,
    force: bool = False,
) -> Path:
    """Build and atomically publish the workbook, refusing silent overwrite."""

    summary, dispatch_rows = load_run_directory(input_directory)
    workbook = build_workbook(summary, dispatch_rows)
    output = Path(output_path).resolve()
    if output.exists() and not force:
        raise FileExistsError(f"refusing to overwrite {output.name}; pass --force to replace it")
    buffer = BytesIO()
    workbook.save(buffer)
    _publish_bytes(output, buffer.getvalue())
    return output
