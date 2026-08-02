from __future__ import annotations

import json
import shutil
import sys
import tempfile
import unittest
from contextlib import redirect_stdout
from importlib.util import find_spec
from io import StringIO
from pathlib import Path
from typing import Any
from unittest.mock import patch

from pv_bess.cli import main
from pv_bess.dispatch import optimize_dispatch
from pv_bess.finance import evaluate_financials
from pv_bess.io import load_scenario, write_results

_OPENPYXL_AVAILABLE = find_spec("openpyxl") is not None
_SKIP_REASON = "the optional Excel dependency is not installed"


class ExportXlsxCommandTests(unittest.TestCase):
    run_directory: Path
    summary: dict[str, Any]

    @classmethod
    def setUpClass(cls) -> None:
        super().setUpClass()
        holder = tempfile.TemporaryDirectory()
        cls.addClassCleanup(holder.cleanup)
        cls.run_directory = Path(holder.name)
        scenario, assumptions = load_scenario(
            Path(__file__).resolve().parents[1] / "sample-data" / "scenario.json"
        )
        dispatch = optimize_dispatch(scenario)
        financial = evaluate_financials(dispatch, scenario, assumptions)
        summary_path, _ = write_results(cls.run_directory, dispatch, financial)
        cls.summary = json.loads(summary_path.read_text(encoding="utf-8"))

    def _export(self, source: Path, target: Path, *extra: str) -> int:
        with redirect_stdout(StringIO()):
            return main(["export-xlsx", "--input", str(source), "--output", str(target), *extra])

    @staticmethod
    def _summary_items(sheet: Any) -> dict[str, tuple[Any, Any]]:
        return {
            row[0]: (row[1], row[2])
            for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True)
            if row[0] is not None
        }

    @unittest.skipUnless(_OPENPYXL_AVAILABLE, _SKIP_REASON)
    def test_export_writes_expected_sheets_and_key_cells(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as directory:
            target = Path(directory) / "report.xlsx"
            self.assertEqual(self._export(self.run_directory, target), 0)
            workbook = load_workbook(target)
            self.assertEqual(workbook.sheetnames, ["Summary", "Cash flows", "Dispatch"])

            summary_sheet = workbook["Summary"]
            self.assertEqual(summary_sheet["A1"].value, "Item")
            self.assertEqual(summary_sheet.freeze_panes, "A2")
            items = self._summary_items(summary_sheet)
            self.assertEqual(items["Scenario name"][0], self.summary["scenario_name"])
            self.assertEqual(
                items["Dispatch input SHA-256"][0], self.summary["dispatch_input_sha256"]
            )
            self.assertEqual(
                items["Analysis input SHA-256"][0], self.summary["analysis_input_sha256"]
            )
            npv_value, npv_unit = items["Net present value"]
            self.assertAlmostEqual(npv_value, -3_818_737.07961859)
            self.assertEqual(npv_unit, "EUR")
            lcos_value, lcos_unit = items["Levelized cost of storage"]
            self.assertAlmostEqual(lcos_value, 265.9735362230494)
            self.assertEqual(lcos_unit, "EUR/MWh")
            self.assertEqual(items["Economic phase status"][0], "optimal")
            column_a = [row[0].value for row in summary_sheet.iter_rows(min_col=1, max_col=1)]
            for limitation in self.summary["limitations"]:
                self.assertIn(limitation, column_a)

            cash_sheet = workbook["Cash flows"]
            cash_flows = self.summary["financial_summary"]["cash_flows_eur"]
            self.assertEqual(cash_sheet["A1"].value, "Year")
            self.assertEqual(cash_sheet["B1"].value, "Cash flow (EUR)")
            self.assertEqual(cash_sheet.freeze_panes, "A2")
            self.assertEqual(cash_sheet.max_row, len(cash_flows) + 1)
            self.assertEqual(cash_sheet["A2"].value, 0)
            self.assertAlmostEqual(cash_sheet["B2"].value, cash_flows[0])
            last = cash_sheet.cell(row=len(cash_flows) + 1, column=2).value
            self.assertAlmostEqual(last, cash_flows[-1])

            dispatch_sheet = workbook["Dispatch"]
            self.assertEqual(dispatch_sheet["A1"].value, "Timestamp")
            self.assertEqual(dispatch_sheet.freeze_panes, "A2")
            self.assertEqual(dispatch_sheet.max_row, 25)
            self.assertEqual(dispatch_sheet["A2"].value, "2026-06-15T00:00:00+00:00")
            self.assertEqual(dispatch_sheet["B2"].value, 1.0)
            self.assertEqual(dispatch_sheet["C2"].value, 0.0)
            self.assertEqual(dispatch_sheet["D2"].value, 45.0)

    @unittest.skipUnless(_OPENPYXL_AVAILABLE, _SKIP_REASON)
    def test_workbook_metadata_stays_anonymous(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as directory:
            target = Path(directory) / "report.xlsx"
            self.assertEqual(self._export(self.run_directory, target), 0)
            properties = load_workbook(target).properties
            self.assertEqual(properties.creator, "pv-bess-hybrid")
            self.assertIn(properties.lastModifiedBy, (None, "pv-bess-hybrid"))
            for attribute in ("title", "subject", "description", "keywords", "category"):
                self.assertIn(getattr(properties, attribute), (None, ""))

    @unittest.skipUnless(_OPENPYXL_AVAILABLE, _SKIP_REASON)
    def test_export_refuses_overwrite_without_force(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            target = Path(directory) / "report.xlsx"
            self.assertEqual(self._export(self.run_directory, target), 0)
            with self.assertRaisesRegex(SystemExit, "refusing to overwrite"):
                self._export(self.run_directory, target)
            self.assertEqual(self._export(self.run_directory, target, "--force"), 0)

    @unittest.skipUnless(_OPENPYXL_AVAILABLE, _SKIP_REASON)
    def test_absent_metrics_render_as_not_available(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as directory:
            run_copy = Path(directory) / "run"
            shutil.copytree(self.run_directory, run_copy)
            summary_path = run_copy / "summary.json"
            payload = json.loads(summary_path.read_text(encoding="utf-8"))
            for key in ("irr_fraction", "lcos_eur_per_mwh"):
                payload["financial_summary"][key] = None
            payload["dispatch_summary"]["curtailment_reduction_fraction"] = None
            payload["solver"]["backend_version"] = None
            for phase in payload["solver"]["phases"].values():
                phase["achieved_relative_mip_gap"] = None
            summary_path.write_text(json.dumps(payload), encoding="utf-8")

            target = Path(directory) / "report.xlsx"
            self.assertEqual(self._export(run_copy, target), 0)
            items = self._summary_items(load_workbook(target)["Summary"])
            self.assertEqual(items["Internal rate of return"][0], "n/a")
            self.assertEqual(items["Levelized cost of storage"][0], "n/a")
            self.assertEqual(items["Curtailment reduction"][0], "n/a")
            self.assertEqual(items["Economic phase achieved MIP gap"][0], "n/a")
            self.assertTrue(str(items["Backend"][0]).endswith("unknown"))

    @unittest.skipUnless(_OPENPYXL_AVAILABLE, _SKIP_REASON)
    def test_export_rejects_mismatched_provenance_hashes(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            run_copy = Path(directory) / "run"
            shutil.copytree(self.run_directory, run_copy)
            dispatch_path = run_copy / "dispatch.csv"
            tampered = dispatch_path.read_text(encoding="utf-8").replace(
                self.summary["dispatch_input_sha256"], "0" * 64
            )
            dispatch_path.write_text(tampered, encoding="utf-8")
            with self.assertRaisesRegex(SystemExit, "does not match the summary.json"):
                self._export(run_copy, Path(directory) / "report.xlsx")

    @unittest.skipUnless(_OPENPYXL_AVAILABLE, _SKIP_REASON)
    def test_export_rejects_malformed_dispatch_files(self) -> None:
        header_line, first_row = (
            (self.run_directory / "dispatch.csv").read_text(encoding="utf-8").splitlines()[:2]
        )
        fields = first_row.split(",")
        fields[3] = "not-a-number"
        cases = (
            ("a,b\n1,2\n", "header must be exactly"),
            (header_line + "\n", "contains no interval rows"),
            (header_line + "\n" + ",".join(fields) + "\n", "invalid dispatch.csv row 2"),
        )
        for content, message in cases:
            with self.subTest(message=message), tempfile.TemporaryDirectory() as directory:
                run_copy = Path(directory) / "run"
                shutil.copytree(self.run_directory, run_copy)
                (run_copy / "dispatch.csv").write_text(content, encoding="utf-8")
                with self.assertRaisesRegex(SystemExit, message):
                    self._export(run_copy, Path(directory) / "report.xlsx")

    @unittest.skipUnless(_OPENPYXL_AVAILABLE, _SKIP_REASON)
    def test_export_rejects_an_incomplete_summary(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            run_copy = Path(directory) / "run"
            shutil.copytree(self.run_directory, run_copy)
            summary_path = run_copy / "summary.json"
            payload = json.loads(summary_path.read_text(encoding="utf-8"))
            del payload["solver"]
            summary_path.write_text(json.dumps(payload), encoding="utf-8")
            with self.assertRaisesRegex(SystemExit, "not a complete run summary"):
                self._export(run_copy, Path(directory) / "report.xlsx")

    @unittest.skipUnless(_OPENPYXL_AVAILABLE, _SKIP_REASON)
    def test_export_reports_a_missing_run_directory(self) -> None:
        with (
            tempfile.TemporaryDirectory() as directory,
            self.assertRaisesRegex(SystemExit, "error: cannot access"),
        ):
            self._export(Path(directory) / "absent", Path(directory) / "report.xlsx")

    def test_export_without_the_xlsx_extra_reports_a_clear_error(self) -> None:
        with (
            patch.dict(sys.modules, {"pv_bess.xlsx": None, "openpyxl": None}),
            self.assertRaisesRegex(SystemExit, r"pip install 'pv-bess-hybrid\[xlsx\]'"),
        ):
            main(
                [
                    "export-xlsx",
                    "--input",
                    str(self.run_directory),
                    "--output",
                    str(self.run_directory / "report.xlsx"),
                ]
            )


if __name__ == "__main__":  # pragma: no cover
    unittest.main()
